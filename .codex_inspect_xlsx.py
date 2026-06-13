from openpyxl import load_workbook
for path in [r"D:\downloads\38833FF26BA1D.UnigramPreview_g9c9v27vpyspw!App\(1)30-05-2026 DOUGLAS GABRIEL.xlsx", r"D:\downloads\38833FF26BA1D.UnigramPreview_g9c9v27vpyspw!App\(2)30-05-2026 DOUGLAS GABRIEL.xlsx"]:
    print('FILE', path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print('SHEET', ws.title, 'ROWS', ws.max_row, 'COLS', ws.max_column)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        print('ROW', i, row)
    print('---')
